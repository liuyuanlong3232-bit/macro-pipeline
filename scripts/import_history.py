#!/usr/bin/env python3
"""导入 D:\commodity_research_platform\export\merged 历史数据到 hermes.db"""
import openpyxl, sqlite3
from pathlib import Path
from datetime import datetime

HISTORY = Path.home() / "hermes-macro-data" / "history"
DB = Path.home() / "hermes-macro-data" / "hermes.db"

conn = sqlite3.connect(str(DB))

files = [
    ("宏观经济.xlsx", "macro"),
    ("商品价格.xlsx", "price"),
    ("能源.xlsx", "energy"),
    ("农业.xlsx", "agri"),
    ("期货持仓.xlsx", "cot"),
    ("新闻.xlsx", "news"),
    ("数据概览.xlsx", "summary"),
]

total_rows = 0
for fname, prefix in files:
    path = HISTORY / fname
    if not path.exists():
        print(f"  ⚠️ {fname} 不存在")
        continue
    wb = openpyxl.load_workbook(str(path), read_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        cols = rows[0]
        data = rows[1:]
        
        # 清理列名：去掉无效字符
        clean_cols = [str(c).replace(" ", "_").replace("-", "_").replace("/", "_")
                      .replace("%", "pct").replace("(", "").replace(")", "")
                      .replace(".", "_").replace("__", "_").rstrip("_")[:64]
                      for c in cols]
        
        table = f"{prefix}_{sname.replace(' ', '_').replace('-', '_')[:40]}"
        
        # 删旧建新
        conn.execute(f'DROP TABLE IF EXISTS "{table}"')
        col_defs = ", ".join([f'"{c}" TEXT' for c in clean_cols])
        conn.execute(f'CREATE TABLE "{table}" ({col_defs})')
        
        # 导入
        q = ",".join(["?"] * len(clean_cols))
        c = ",".join([f'"{x}"' for x in clean_cols])
        for row in data:
            vals = [str(v) if v is not None else None for v in row]
            if len(vals) == len(clean_cols):
                conn.execute(f'INSERT INTO "{table}" ({c}) VALUES ({q})', vals)
        
        cnt = conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
        total_rows += cnt
        print(f"  ✅ {fname}/{sname} → {table} ({cnt}行)")

conn.commit()
conn.close()
print(f"\n📦 总计: {total_rows}行历史数据已导入")
